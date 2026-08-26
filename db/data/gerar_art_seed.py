# -*- coding: utf-8 -*-
"""
Obras de arte (DMG) → seed revisável.

Kind PRÓPRIO (`treasure`), não matéria-prima: obra de arte não se consome para
fabricar nada — carrega-se, vende-se e distribui-se como saque.

⚠️ A planilha dá o peso em LB; o banco é canônico em KG. O fator é o do LIVRO
(2.0, `EquipmentRules::LB_PER_KG`), não o físico 2.20462 — com o físico os
números do livro nunca batem. Gravar 2.0 no campo kg faria a obra pesar o dobro.
"""
import openpyxl, json, re, sys, unicodedata, collections

S = "/private/tmp/claude-501/-Users-alyssonjosesoares-Documents-lafiga-front-lafiga/3d48da8b-c253-4475-b774-cebbdbaf6263/scratchpad/arte"
LB_PER_KG = 2.0

# Faixas do DMG: o mestre rola DENTRO de uma faixa para gerar o saque.
FAIXAS = {25: "25 po", 250: "250 po", 750: "750 po", 2500: "2500 po", 7500: "7500 po"}

def slug(s):
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", s.lower())).strip("-")

ws = openpyxl.load_workbook("/Users/alyssonjosesoares/Downloads/Tabelas de Obras de arte.xlsx",
                            data_only=True)["Página1"]
itens, relat, vistos = [], collections.defaultdict(list), {}
for r in range(3, ws.max_row + 1):
    nome = ws.cell(row=r, column=1).value
    if nome is None or not str(nome).strip(): continue
    nome = str(nome).strip()
    preco = ws.cell(row=r, column=2).value
    peso_lb = ws.cell(row=r, column=3).value

    k = slug(nome)
    if k in vistos:
        relat["NOME DUPLICADO (aborta)"].append(f"{nome} × linha {vistos[k]}")
        continue
    vistos[k] = r

    if preco is None:
        relat["preço ausente (aborta)"].append(nome); continue
    preco = float(preco)
    if int(preco) not in FAIXAS:
        relat["faixa DESCONHECIDA (aborta)"].append(f"{nome}: {preco}"); continue

    if peso_lb is None:
        relat["peso ausente (needs_review)"].append(nome)
        peso_kg = None
    else:
        peso_kg = round(float(peso_lb) / LB_PER_KG, 2)

    itens.append({
        "api_index": f"art-{slug(nome)}",
        "name": nome,
        "kind": "treasure",
        "category": "art",
        "value_gp": preco,
        # KG no banco; a fronteira ×2 devolve o lb do livro na serialização.
        "weight_kg": peso_kg,
        "source": "DMG — Objetos de Arte",
        "props": {
            "unit": "un",
            # A faixa é o que o DMG usa para sortear saque; guardada explícita
            # para o filtro não ter de re-derivar do preço.
            "art_tier": int(preco),
        },
    })

aborta = {k: v for k, v in relat.items() if "aborta" in k and v}
print(f"linhas → itens: {len(itens)}")
for cat, lst in sorted(relat.items()):
    print(f"\n── {cat}: {len(lst)}")
    for x in lst[:8]: print(f"     {x}")
if aborta:
    print("\n❌ ABORTADO"); sys.exit(1)

por_faixa = collections.Counter(i["props"]["art_tier"] for i in itens)
print("por faixa:", dict(sorted(por_faixa.items())))
print("peso kg:", dict(sorted(collections.Counter(i["weight_kg"] for i in itens).items())))
json.dump(itens, open(S + "/art_seed.json", "w"), ensure_ascii=False, indent=1)
print(f"\n✅ art_seed.json com {len(itens)} obras")
