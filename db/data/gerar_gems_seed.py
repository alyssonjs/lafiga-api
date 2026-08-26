# -*- coding: utf-8 -*-
"""
Gemas (homebrew) → seed revisável.

⚠️ A planilha tem DUAS unidades: as linhas de gema dizem "10PO" e a tabela de
balanceamento diz "10 PL". Uma peça de platina vale 10 po — se a tabela
estivesse certa, toda gema valeria 10x mais. A LINHA DA GEMA manda: é o dado
por item, e a tabela é um resumo de faixas. Registrado no relatório.
"""
import openpyxl, json, re, sys, unicodedata, collections

S = "/private/tmp/claude-501/-Users-alyssonjosesoares-Documents-lafiga-front-lafiga/3d48da8b-c253-4475-b774-cebbdbaf6263/scratchpad/gemas"
TIERS = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6}

def slug(s):
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", s.lower())).strip("-")

ws = openpyxl.load_workbook("/Users/alyssonjosesoares/Downloads/Gemas.xlsx", data_only=True)["Gemas"]
itens, relat, vistos = [], collections.defaultdict(list), {}

for r in range(3, ws.max_row + 1):
    nome = ws.cell(row=r, column=2).value
    if nome is None or not str(nome).strip(): continue
    nome = str(nome).strip()
    if nome == "Nome": continue          # cabeçalho

    k = slug(nome)
    if k in vistos:
        relat["NOME DUPLICADO (aborta)"].append(f"{nome} × linha {vistos[k]}"); continue
    vistos[k] = r

    tier_raw = str(ws.cell(row=r, column=7).value or "").strip()
    if tier_raw not in TIERS:
        relat["tier DESCONHECIDO (aborta)"].append(f"{nome}: {tier_raw!r}"); continue

    valor_raw = str(ws.cell(row=r, column=8).value or "").strip()
    m = re.match(r"^([\d.]+)\s*(PO|PL|PP|PC)$", valor_raw, re.I)
    if not m:
        relat["valor NÃO PARSEADO (aborta)"].append(f"{nome}: {valor_raw!r}"); continue
    valor = float(m.group(1).replace(".", ""))
    if m.group(2).upper() != "PO":
        relat["valor fora de PO (revisar)"].append(f"{nome}: {valor_raw}")

    def txt(col):
        v = ws.cell(row=r, column=col).value
        return str(v).strip() if v is not None and str(v).strip() else None

    arma = txt(5)
    if not arma:
        relat["sem efeito de ARMA (fica só vestuário)"].append(nome)

    itens.append({
        "api_index": f"gem-{k}",
        "name": nome,
        "kind": "material",
        "category": "gem",
        "value_gp": valor,
        "source": "Gemas de Lafiga (homebrew)",
        "description": txt(3),                  # aparência é a descrição visual
        "props": {
            "unit": "un",
            "gem_tier": TIERS[tier_raw],
            "gem_power": txt(4),                # poder místico
            "gem_weapon_effect": arma,          # encaixada em arma
            "gem_apparel_effect": txt(6),       # encaixada em vestuário
        },
    })

aborta = {k: v for k, v in relat.items() if "aborta" in k and v}
print(f"gemas: {len(itens)}")
for cat, lst in sorted(relat.items()):
    print(f"\n── {cat}: {len(lst)}")
    for x in lst[:6]: print(f"     {x}")
if aborta:
    print("\n❌ ABORTADO"); sys.exit(1)

print("\npor tier:", dict(sorted(collections.Counter(i['props']['gem_tier'] for i in itens).items())))
print("por valor:", dict(sorted(collections.Counter(i['value_gp'] for i in itens).items())))
json.dump(itens, open(S + "/gems_seed.json", "w"), ensure_ascii=False, indent=1)
print(f"\n✅ gems_seed.json com {len(itens)} gemas")
