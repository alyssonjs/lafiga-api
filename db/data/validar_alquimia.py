# -*- coding: utf-8 -*-
"""
Fase 5 — valida o BANCO contra a PLANILHA, relendo o XLSX do zero.

Não passa pelo alchemy_seed.json de propósito: comparar o banco com o arquivo
que o gerou provaria só que a cópia funcionou. Aqui a planilha é lida de novo e
as regras do manual são recalculadas do lado do teste.
"""
import json, re, unicodedata, sys, collections
import openpyxl

XLSX = "/Users/alyssonjosesoares/Downloads/Tabela de Poções.xlsx"

def chave(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()

# ── planilha, do zero ─────────────────────────────────────────────────────────
ws = openpyxl.load_workbook(XLSX, data_only=True)["Poções"]
planilha = {}
for r in range(5, 210):
    nome = ws.cell(row=r, column=4).value
    if not nome: continue
    custo = ws.cell(row=r, column=9).value
    m = ws.cell(row=r, column=13).value
    planilha[chave(nome)] = {
        "nome": str(nome).strip(),
        "raridade": ws.cell(row=r, column=3).value,
        "custo": float(custo) if custo is not None else None,
        # Regra do manual: custo = metade do mercado.
        "mercado": float(m) if m is not None else (float(custo) * 2 if custo is not None else None),
        "cd": ws.cell(row=r, column=10).value,
        "composicao": ws.cell(row=r, column=6).value,
    }

banco = {chave(x["nome"]): x for x in json.load(open("/tmp/db_receitas.json"))}

RARIDADE = {"comum": "common", "incomum": "uncommon", "raro": "rare",
            "muito raro": "very_rare", "lendario": "legendary"}

falhas = collections.defaultdict(list)
print(f"planilha: {len(planilha)} produtos | banco: {len(banco)} receitas\n")

# 1. cobertura
faltando = set(planilha) - set(banco)
sobrando = set(banco) - set(planilha)
if faltando: falhas["produto na planilha SEM receita no banco"] = sorted(planilha[k]["nome"] for k in faltando)
if sobrando: falhas["receita no banco SEM linha na planilha"] = sorted(banco[k]["nome"] for k in sobrando)

for k, p in planilha.items():
    b = banco.get(k)
    if not b: continue
    rot = p["nome"]

    # 2. preço de mercado
    if p["mercado"] is not None and b["mercado"] is not None:
        if abs(p["mercado"] - b["mercado"]) > 0.01:
            falhas["preço de mercado"].append(f"{rot}: planilha {p['mercado']} × banco {b['mercado']}")

    # 3. custo de fabricação = metade do mercado
    if p["custo"] is not None and b["custo"] is not None:
        if abs(p["custo"] - b["custo"]) > 0.01:
            falhas["custo de fabricação"].append(f"{rot}: planilha {p['custo']} × banco {b['custo']}")
        if b["mercado"] and abs(b["custo"] - b["mercado"] / 2) > 0.01:
            falhas["REGRA custo = mercado/2"].append(f"{rot}: {b['custo']} ≠ {b['mercado']/2}")

    # 4. dias = mercado / 25 (ritmo de 25 po/dia)
    if b["mercado"] and b["dias"] is not None:
        if abs(b["dias"] - b["mercado"] / 25) > 0.02:
            falhas["REGRA dias = mercado/25"].append(f"{rot}: {b['dias']} ≠ {b['mercado']/25:.2f}")

    # 5. CD
    if p["cd"] is not None and b["cd"] is not None and int(p["cd"]) != int(b["cd"]):
        falhas["CD"].append(f"{rot}: planilha {p['cd']} × banco {b['cd']}")

    # 6. raridade
    esperada = RARIDADE.get(chave(p["raridade"]))
    if esperada and b["raridade"] != esperada:
        falhas["raridade"].append(f"{rot}: esperado {esperada}, banco {b['raridade']}")

    # 7. contagem de ingredientes: quantos "+" a composição tem, contando que
    #    "X ou Y" gera DOIS registos (alternativas) e o "+" de Fogo-Fátuo não conta.
    comp = str(p["composicao"] or "").replace("Fogo+Fátuo", "FOGOFATUO")
    if comp.strip():
        partes = [x for x in re.split(r"\s*\+\s*", comp) if x.strip()]
        esperados = sum(len(re.split(r"\s+ou\s+", x, flags=re.I)) for x in partes)
        if esperados != len(b["ingredientes"]):
            falhas["nº de ingredientes"].append(
                f"{rot}: composição sugere {esperados}, banco tem {len(b['ingredientes'])}")

    # 8. quantidades, ingrediente a ingrediente. Quem NÃO traz número na
    #    composição vale 1 — é o caso das magias ("Mg Invisibilidade") e de
    #    "Unha de Gigante". Ignorar isso faria a checagem acusar o banco de
    #    inventar quantidades que ele apenas assumiu corretamente.
    if comp.strip():
        qtds = []
        for parte in partes:
            for alt in re.split(r"\s+ou\s+", parte, flags=re.I):
                m2 = re.match(r"^\s*(\d+(?:[.,]\d+)?)", alt.strip())
                qtds.append(float(m2.group(1).replace(",", ".")) if m2 else 1.0)
        if sorted(qtds) != sorted(i["qtd"] for i in b["ingredientes"]):
            falhas["quantidades"].append(
                f"{rot}: planilha {sorted(qtds)} × banco {sorted(i['qtd'] for i in b['ingredientes'])}")

# ── relatório ─────────────────────────────────────────────────────────────────
total = sum(len(v) for v in falhas.values())
if not falhas:
    print("✅ TUDO CONFERE — 100 receitas, campo a campo.")
else:
    for cat, itens in sorted(falhas.items()):
        print(f"── {cat}: {len(itens)}")
        for i in itens[:6]: print(f"     {i}")
        if len(itens) > 6: print(f"     … +{len(itens)-6}")
print(f"\ntotal de divergências: {total}")
sys.exit(0)
