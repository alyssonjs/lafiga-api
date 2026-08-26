# -*- coding: utf-8 -*-
"""
F4 — valida o BANCO contra as DUAS planilhas, relendo os XLSX do zero.

Mesmo desenho do validar_alquimia: não passa pelo herbs_seed.json — comparar o
banco com o arquivo que o gerou provaria só que a cópia funcionou. Os campos
preservados em bruto (descrição, efeito, rendimento, localização original)
conferem por IGUALDADE com a célula; os normalizados (raridade, estação,
chance) são re-derivados aqui com mapas próprios.
"""
import json, re, sys, unicodedata, collections
import openpyxl

def chave(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", s.lower())).strip()

RARIDADE = {"comum": "common", "incomum": "uncommon", "raro": "rare", "rara": "rare",
            "muito raro": "very_rare", "lendario": "legendary", "legendario": "legendary",
            "maravilhoso": None}
ESTACAO = {"primavera": "primavera", "verao": "verao", "veao": "verao", "primavea": "primavera",
           "outono": "outono", "inverno": "inverno", "inveno": "inverno"}
PREPARO = {"maceracao": "Maceração", "maceracao seca": "Maceração Seca", "cataplasma": "Cataplasma",
           "banho": "Banho", "destilado": "Destilado", "oleo": "Óleo", "cozimento": "Cozimento",
           "perfume": "Perfume", "po": "Pó vegetal", "po vegetal": "Pó vegetal",
           "pos vegetais": "Pó vegetal", "cha por decoccao": "Chá por decocção",
           "infusao fria": "Infusão fria", "cha por infusao fria": "Infusão fria",
           "infusao quente": "Infusão quente", "cha por infusao quente": "Infusão quente"}
USO = {"ingestao": ["ingestao"], "contato": ["contato"], "inalacao": ["inalacao"],
       "ferimento": ["ferimento"], "ingestao contado": ["ingestao", "contato"]}
FORCADOS_DMG = {"muco de verme da carnica": "Muco de Rastejante",
                "serum da verdade": "Soro da Verdade",
                "vapores causticantes de othur": "Fumaça de Othur Queimado",
                "essencia de eter": "Essência de Éter", "tintura palida": "Tintura Pálida",
                "veneno de serpente": "Veneno de Serpente",
                "veneno de verme purpura": "Veneno de Verme Púrpura",
                "veneno de wyvern": "Veneno de Wyvern", "veneno drow": "Veneno Drow"}
VEGETAIS = {"planta", "plata", "fungo", "arvore", "raiz"}
# Fora das planilhas, mas legitimamente na família: veio da ALQUIMIA — e por
# isso o prefixo é `mat-`, não `herb-` (cada importador assina o seu).
EXCECOES_BANCO = {"mat-polpa-do-kiuvi"}

def estacoes_de(raw):
    if raw is None or str(raw).strip() in ("-", ""): return [], None
    txt = str(raw).strip()
    if chave(txt) == "todas as estacoes":
        return ["primavera", "verao", "outono", "inverno"], None
    nota = None; base = txt
    m = re.match(r"^(durante a noite de|noites de lua nova do|noites de|inicio da)\s+(.*)$", chave(txt))
    if m: nota, base = txt, m.group(2)
    lista = []
    for p in re.split(r"/| e ", str(base)):
        k = chave(p)
        if k in ESTACAO and ESTACAO[k] not in lista: lista.append(ESTACAO[k])
    return lista, nota

def valor_de(raw):
    if raw is None or not str(raw).strip(): return None
    m = re.match(r"^\s*([\d.,]+)\s*p?o?\s*$", str(raw), re.I)
    return float(m.group(1).replace(",", ".")) if m else None

def chance_de(raw, estacoes):
    if raw is None or str(raw).strip() in ("-", ""): return None, None
    txt = str(raw).strip()
    pares = re.findall(r"([\d.,]+)\s*%", txt)
    if len(pares) >= 2:
        vals = [float(p.replace(",", ".")) / 100 for p in pares]
        if len(vals) == len(estacoes): return None, dict(zip(estacoes, vals))
        return max(vals), None
    try: v = float(txt.replace("%", "").replace(",", "."))
    except ValueError: return None, None
    return (v / 100 if v > 1 else v), None

# ── planilhas, do zero ────────────────────────────────────────────────────────
esperado, dmg_esperado = {}, {}
contagem_nome = collections.Counter()
for arq, familia in [("/Users/alyssonjosesoares/Downloads/Tabela de Ervas (Herbalismo).xlsx", "herb"),
                     ("/Users/alyssonjosesoares/Downloads/Tabela de Venenos.xlsx", "poison-herb")]:
    ws = openpyxl.load_workbook(arq, data_only=True)[list(openpyxl.load_workbook(arq).sheetnames)[0]]
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(row=r, column=1).value
        if nome is None or not str(nome).strip(): continue
        l = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        nome = str(nome).strip()
        k = chave(nome)
        eh_dmg = familia == "poison-herb" and (chave(l[6]) not in VEGETAIS or k in FORCADOS_DMG)
        estacoes, nota = estacoes_de(l[12])
        reg = {
            "nome": nome, "familia": familia,
            "descricao": (str(l[1]).strip() if l[1] else None),
            "efeito": (str(l[2]).strip() if l[2] else None),
            "preparo": PREPARO.get(chave(l[3])), "uso": USO.get(chave(l[14])),
            "valor": valor_de(l[4]), "raridade": RARIDADE.get(chave(l[8]), "?!"),
            "cd": int(l[9]) if l[9] is not None else None,
            "rendimento": (str(l[10]).strip() if l[10] else None),
            "local_raw": (str(l[7]).strip() if l[7] and str(l[7]).strip() != "-" else None),
            "estacoes": estacoes, "nota_estacao": nota,
        }
        reg["chance"], reg["chance_por_estacao"] = chance_de(l[13], estacoes)
        if eh_dmg:
            dmg_esperado[FORCADOS_DMG.get(k, nome)] = reg
        else:
            contagem_nome[(familia, k)] += 1
            n = contagem_nome[(familia, k)]
            reg["variante"] = n
            esperado[(familia, k, n)] = reg

banco = json.load(open("/tmp/db_ervas.json"))
falhas = collections.defaultdict(list)

# ── casa o banco com a planilha ───────────────────────────────────────────────
def slug(s):
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", s.lower())).strip("-")

vistos_banco = set()
for it in banco["materiais"]:
    if it["api_index"] in EXCECOES_BANCO: continue
    fam, k = it["category"], chave(it["name"])
    m = re.match(r"^(?:herb|vherb)-(.*?)(?:-(\d+))?$", it["api_index"])
    variante = int(m.group(2)) if m and m.group(2) else 1
    ref = esperado.get((fam, k, variante))
    if ref is None:
        falhas["item no banco SEM linha na planilha"].append(it["api_index"]); continue
    vistos_banco.add((fam, k, variante))
    rot = f"{it['name']} v{variante}" if contagem_nome[(fam, k)] > 1 else it["name"]
    f, p = it.get("foraging") or {}, it.get("preparation") or {}

    def cmp(campo, a, b):
        if a != b: falhas[campo].append(f"{rot}: planilha {a!r} × banco {b!r}")

    cmp("valor", ref["valor"], it["value_gp"])
    cmp("raridade", ref["raridade"], it["rarity"])
    cmp("descrição (bruto)", ref["descricao"], it["description"])
    cmp("efeito (bruto)", ref["efeito"], p.get("effect"))
    cmp("preparo", ref["preparo"], p.get("method"))
    cmp("uso", ref["uso"], p.get("usage"))
    cmp("CD", ref["cd"], f.get("dc"))
    cmp("rendimento (bruto)", ref["rendimento"], f.get("yield"))
    cmp("localização original (bruto)", ref["local_raw"], f.get("locations_raw"))
    cmp("estações", ref["estacoes"], f.get("seasons") or [])
    cmp("nota de estação", ref["nota_estacao"], f.get("season_note"))
    cmp("chance", ref["chance"], f.get("chance"))
    cmp("chance por estação", ref["chance_por_estacao"], f.get("chance_by_season"))

faltando = set(esperado) - vistos_banco
for fam, k, n in sorted(faltando):
    falhas["linha da planilha SEM item no banco"].append(f"{esperado[(fam,k,n)]['nome']} ({fam} v{n})")

# ── venenos DMG enriquecidos ──────────────────────────────────────────────────
banco_dmg = {x["name"]: x for x in banco["venenos_dmg"]}
for nome_cat, ref in dmg_esperado.items():
    alvo = banco_dmg.get(nome_cat)
    if alvo is None:
        falhas["veneno DMG sem enriquecimento"].append(nome_cat); continue
    f, p = alvo.get("foraging") or {}, alvo.get("preparation") or {}
    if f.get("dc") != ref["cd"]:
        falhas["DMG: CD de extração"].append(f"{nome_cat}: {ref['cd']} × {f.get('dc')}")
    if p.get("method") != ref["preparo"]:
        falhas["DMG: preparo"].append(f"{nome_cat}: {ref['preparo']} × {p.get('method')}")
    if p.get("effect") != ref["efeito"]:
        falhas["DMG: efeito (bruto)"].append(f"{nome_cat}: difere")
sobra_dmg = set(banco_dmg) - set(dmg_esperado)
if sobra_dmg: falhas["enriquecimento sem linha na planilha"] = sorted(sobra_dmg)

# ── relatório ─────────────────────────────────────────────────────────────────
total = sum(len(v) for v in falhas.values())
print(f"planilhas: {len(esperado)} materiais + {len(dmg_esperado)} DMG | banco: {len(banco['materiais'])-len(EXCECOES_BANCO)} + {len(banco['venenos_dmg'])}")
if not falhas:
    print("\n✅ TUDO CONFERE — campo a campo, incluindo os textos brutos.")
else:
    for cat, itens in sorted(falhas.items()):
        print(f"\n── {cat}: {len(itens)}")
        for x in itens[:6]: print(f"     {x}")
        if len(itens) > 6: print(f"     … +{len(itens)-6}")
print(f"\ntotal de divergências: {total}")
