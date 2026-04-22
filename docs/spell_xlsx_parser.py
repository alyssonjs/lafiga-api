#!/usr/bin/env python3
"""
Parses api/docs/spelldatabase.xlsx (PHB pt-BR spell database) and emits a
structured JSON file `api/docs/spelldatabase.parsed.json` consumed by the
rake task `spells:import_xlsx` to enrich api/config/spells.yml.

Usage (from repo root):
    python3 api/docs/spell_xlsx_parser.py

Outputs:
    api/docs/spelldatabase.parsed.json    -- list of {name, level, school, ...}
    api/docs/spelldatabase.parsed.skipped.txt -- index rows we skipped, one per line

The script is purely a syntactic transform (xlsx cell -> structured JSON).
It does not modify the xlsx; it only reads from it.
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("error: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = REPO_ROOT / "api" / "docs" / "spelldatabase.xlsx"
JSON_PATH = REPO_ROOT / "api" / "docs" / "spelldatabase.parsed.json"
SKIPPED_PATH = REPO_ROOT / "api" / "docs" / "spelldatabase.parsed.skipped.txt"

SCHOOL_PT_TO_EN = {
    "abjuracao": "Abjuration",
    "adivinhacao": "Divination",
    "conjuracao": "Conjuration",
    "encantamento": "Enchantment",
    "evocacao": "Evocation",
    "ilusao": "Illusion",
    "necromancia": "Necromancy",
    "transmutacao": "Transmutation",
}

# Header line variants:
#   "2° nível de encantamento"
#   "2º nível de encantamento"
#   "4° nível de adivinhação (ritual)"
#   "Truque de Transmutação"
#   "Truque de transmutacao"
HEADER_RE = re.compile(
    r"^\s*(?:(\d+)\s*[º°]\s*n[íi]vel\s+de\s+([A-Za-zçÇãÁÉÍÓÚáéíóúâêôõ]+)|"
    r"truque\s+de\s+([A-Za-zçÇãÁÉÍÓÚáéíóúâêôõ]+))\s*(\(ritual\))?",
    re.IGNORECASE,
)

FIELD_PATTERNS = OrderedDict(
    [
        ("casting_time", re.compile(r"^\s*Tempo de Conjura[çc][ãa]o\s*:\s*(.+)$", re.IGNORECASE)),
        ("range", re.compile(r"^\s*Alcance\s*:\s*(.+)$", re.IGNORECASE)),
        ("components_raw", re.compile(r"^\s*Componentes\s*:\s*(.+)$", re.IGNORECASE)),
        ("duration", re.compile(r"^\s*Dura[çc][ãa]o\s*:\s*(.+)$", re.IGNORECASE)),
        ("classes_pt", re.compile(r"^\s*(?:Conjuradores|Classes?)\s*:\s*(.+)$", re.IGNORECASE)),
    ]
)

CONTINUATION_FIELDS = {"casting_time", "range", "components_raw", "duration", "classes_pt"}

HIGHER_RE = re.compile(r"^\s*em\s+n[íi]ve(?:is|l)\s+superior(?:es)?\.?\s*", re.IGNORECASE)


def fold(s: str) -> str:
    """Lowercase + strip diacritics for case-insensitive comparisons."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower()) if unicodedata.category(c) != "Mn"
    ).strip()


def parse_components(raw: str) -> tuple[list[str], str | None]:
    raw = raw.strip()
    material = None
    mat = re.search(r"M\s*\(([^)]*)\)", raw, re.IGNORECASE)
    if mat:
        material = re.sub(r"\s+", " ", mat.group(1)).strip()
    parts: list[str] = []
    if re.search(r"\bV\b", raw, re.IGNORECASE):
        parts.append("V")
    if re.search(r"\bS\b", raw, re.IGNORECASE):
        parts.append("S")
    if re.search(r"\bM\b", raw, re.IGNORECASE):
        parts.append("M")
    return parts, material


def find_header_index(lines: list[str]) -> int:
    """Return index of the first line matching HEADER_RE, scanning up to 5 non-empty lines.
    Returns -1 if no header found."""
    seen = 0
    for i, ln in enumerate(lines):
        if not ln.strip():
            continue
        seen += 1
        if seen > 5:
            break
        if HEADER_RE.match(ln):
            return i
    return -1


def join_paragraphs(lines: list[str]) -> list[str]:
    """Collapse soft line wraps into paragraphs separated by blank lines."""
    out: list[str] = []
    cur: list[str] = []
    for x in lines:
        if not x.strip():
            if cur:
                joined = re.sub(r"\s+", " ", " ".join(cur)).strip()
                if joined:
                    out.append(joined)
                cur = []
        else:
            cur.append(x.strip())
    if cur:
        joined = re.sub(r"\s+", " ", " ".join(cur)).strip()
        if joined:
            out.append(joined)
    return out


def parse_spell(name: str, raw_desc: str) -> dict | None:
    if not name or not raw_desc:
        return None
    lines = raw_desc.split("\n")
    h_idx = find_header_index(lines)
    if h_idx < 0:
        return None
    header = lines[h_idx]
    m = HEADER_RE.match(header)
    if not m:
        return None
    level = 0 if m.group(3) else int(m.group(1))
    school_raw = (m.group(2) or m.group(3) or "").strip()
    school_key = fold(school_raw)
    school = SCHOOL_PT_TO_EN.get(school_key, school_raw.title())
    ritual = bool(m.group(4))

    fields: dict[str, str] = {}
    current_field: str | None = None
    body_start = len(lines)

    i = h_idx + 1
    while i < len(lines):
        ln = lines[i]
        stripped = ln.strip()
        if not stripped:
            current_field = None
            i += 1
            continue
        matched_key: str | None = None
        matched_val: str | None = None
        for key, pat in FIELD_PATTERNS.items():
            mm = pat.match(stripped)
            if mm:
                matched_key = key
                matched_val = mm.group(1).strip()
                break
        if matched_key:
            current_field = matched_key
            existing = fields.get(matched_key, "")
            fields[matched_key] = (existing + " " + matched_val).strip() if existing else matched_val
            i += 1
            continue
        if current_field in CONTINUATION_FIELDS and not re.match(
            r"^[A-ZÁÉÍÓÚÂÊÔÃÕÇ]", stripped
        ):
            fields[current_field] = (fields[current_field] + " " + stripped).strip()
            i += 1
            continue
        body_start = i
        break

    body_lines = lines[body_start:]
    desc_lines: list[str] = []
    higher_lines: list[str] = []
    in_higher = False
    for ln in body_lines:
        stripped = ln.strip()
        if not in_higher and stripped and HIGHER_RE.match(stripped):
            in_higher = True
            tail = HIGHER_RE.sub("", stripped).strip()
            if tail:
                higher_lines.append(tail)
            continue
        if in_higher:
            higher_lines.append(ln)
        else:
            desc_lines.append(ln)

    desc_paragraphs = join_paragraphs(desc_lines)
    higher_paragraphs = join_paragraphs(higher_lines)

    components: list[str] = []
    material: str | None = None
    if "components_raw" in fields:
        components, material = parse_components(fields["components_raw"])

    classes_pt: list[str] = []
    if "classes_pt" in fields:
        raw_cls = fields["classes_pt"]
        classes_pt = [
            re.sub(r"\s+", " ", c).strip().lower()
            for c in re.split(r"[,;]", raw_cls)
            if c.strip()
        ]

    duration = re.sub(r"\s+", " ", fields.get("duration", "")).strip()
    casting_time = re.sub(r"\s+", " ", fields.get("casting_time", "")).strip()
    range_str = re.sub(r"\s+", " ", fields.get("range", "")).strip()
    concentration = bool(re.search(r"concentra", duration, re.IGNORECASE))

    return {
        "name": re.sub(r"\s+", " ", name).strip(),
        "level": level,
        "school": school,
        "school_pt": school_key,
        "ritual": ritual,
        "casting_time": casting_time,
        "range": range_str,
        "components": components,
        "material": material,
        "duration": duration,
        "concentration": concentration,
        "desc": desc_paragraphs,
        "higher_level": higher_paragraphs,
        "classes_pt": classes_pt,
    }


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"error: {XLSX_PATH} not found", file=sys.stderr)
        return 1
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    parsed: list[dict] = []
    skipped: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        name = row[0]
        desc = row[1]
        if not name or not desc:
            continue
        result = parse_spell(str(name), str(desc))
        if result:
            parsed.append(result)
        else:
            skipped.append(str(name))

    parsed.sort(key=lambda x: (x["level"], fold(x["name"])))
    JSON_PATH.write_text(json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8")
    SKIPPED_PATH.write_text("\n".join(skipped) + "\n", encoding="utf-8")

    print(f"parsed spells: {len(parsed)} -> {JSON_PATH.relative_to(REPO_ROOT)}")
    print(f"skipped index rows: {len(skipped)} -> {SKIPPED_PATH.relative_to(REPO_ROOT)}")
    by_level: dict[int, int] = {}
    for s in parsed:
        by_level[s["level"]] = by_level.get(s["level"], 0) + 1
    print("by level:", ", ".join(f"L{k}={v}" for k, v in sorted(by_level.items())))
    return 0


if __name__ == "__main__":
    sys.exit(main())
