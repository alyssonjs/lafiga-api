#!/usr/bin/env python3
"""
Extrai as 39 fichas xlsx em api/docs/Ficha de acompanhamento Pre (1).xlsx
para um JSON estruturado normalizado em api/docs/imported_sheets.json.

Estrategia hibrida:
  - Posicoes FIXAS para o que e estavel em 100% das fichas (habilidades,
    pericias, resistencias, HP, CA, deslocamento, sobrecarga, peso, idiomas,
    moedas).
  - Busca por LABEL para o que varia (Folego, HP Extra, Espacos de Magia por
    nivel, Itens Sincronizados, Aljava, Surto de Acao, Indomavel, etc).

Uso:
  python3 api/scripts/extract_xlsx_sheets.py

Saida:
  api/docs/imported_sheets.json
  api/docs/imported_sheets.summary.txt  (relatorio humano)
"""
from __future__ import annotations

import json
import re
import sys
import os
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC_XLSX = ROOT / "api" / "docs" / "Ficha de acompanhamento Pre (1).xlsx"
OUT_JSON = ROOT / "api" / "docs" / "imported_sheets.json"
OUT_TXT = ROOT / "api" / "docs" / "imported_sheets.summary.txt"

ABAS_IGNORAR = {"Medias", "Plan1", "Plan2"}

# Personagens cuja raca nao consta na xlsx — informacao confirmada por humano
# (campanha LaFiga, Abr/2026). Mapeia tab_name -> (race_api_index, subrace_api_index).
# Use None para subraca quando nao especificada (auditor aceita).
MANUAL_RACE_OVERRIDES = {
    "Nayara":     ("human",    "standard"),  # Abigail Le Fay
    "Fininho":    ("elf",      "wood"),      # Adimael (Elfo da Floresta)
    "Caio":       ("gnome",    "rock"),      # Stivi Magal (gnomo)
    "Allan":      ("human",    "standard"),  # Sirius Bastiao de Ferro
    "João":       ("half_elf", None),        # Lenny (meio-elfo)
    "Miguel":     ("human",    "standard"),  # Rolander
    "Alieksey":   ("tiefling", "infernal"),  # Darkmenos
    "Aberrama":   ("tiefling", "infernal"),  # Aberama Gold
    "Angelina":   ("tiefling", "infernal"),  # Angelina
}

# Personagens a ignorar do audit / fixtures (a pedido do GM):
#   - Modelo: ficha-template
#   - Drugoy / Tony Ramos / Pandora: nao serao testados
SKIP_FROM_AUDIT = {"Modelo", "Drugoy", "Tony Ramos", "Pandora"}

# Subclasses nao declaradas na xlsx — defaults razoaveis para destravar
# fixtures de teste (TODO confirmar com GM se algum destes nao bate com
# a campanha real). Mapeia tab_name -> subclass_api_index.
MANUAL_SUBCLASS_OVERRIDES = {
    "Nayara":   "evocacao",   # Mago — escola padrao PHB (api_index do DB)
    "Miguel":   "devotion",   # Paladino — Juramento de Devocao
    "Alieksey": "fiend",      # Bruxo — O Infero (Tiefling -> default)
}

# Antecedente (Background) por aba.
#
# A xlsx NAO tem celula dedicada de antecedente; ate aqui o front inferia o
# bg pelas duas pericias com mais horas fora do pool de classe. Quando uma
# pericia do background tambem esta no pool da classe (ex.: Atletismo no
# Soldado pra um Barbaro), ela era "consumida" pelas picks de classe e o
# match falhava — caindo no Acolito como ultimo recurso.
#
# Mapeia tab_name -> background id no `MOCK_BACKGROUNDS_FOR_INFERENCE`
# (front-lafiga/src/services/importedSheets/importedSheetSkillsFromTraining.ts):
#   bg-1 Acolito | bg-2 Soldado  | bg-3 Sabio   | bg-4 Criminoso
#   bg-5 Nobre   | bg-6 Eremita  | bg-7 Forasteiro
#   bg-8 Artesao de Guilda | bg-9 Charlatao
#
# TODO: ir preenchendo com o GM. Quem nao constar continua no comportamento
# antigo (heuristica + Acolito de fallback).
MANUAL_BACKGROUND_OVERRIDES = {
    "Rorinar": "bg-5",  # Nobre (anao da colina, Barbaro do Caminho do Guerreiro Urso)
}

# ---------------------------------------------------------------------------
# Mapeamentos canonicos (Race / SubRace / Klass / SubKlass)
# ---------------------------------------------------------------------------
# Os api_index abaixo refletem o que existe no projeto:
#   - Race.api_index / SubRace.api_index (db/seeds + RACE_RULES)
#   - ClassRules::CLASS_RULES (api/app/services/class_rules.rb)
# ---------------------------------------------------------------------------

# Nome canonico (em PT, normalizado) -> (race_api, subrace_api or None)
RACE_PATTERNS = [
    # ordem importa: padroes mais especificos primeiro
    (r"\b(humano|humana)\s+argulano\b",        "human",      "variant"),
    (r"\b(humano|humana)\s+kumarense\b",       "human",      "variant"),
    (r"\b(humano|humana)\s+shou\b",            "human",      "variant"),
    (r"\b(humano|humana)\s+variante\b",        "human",      "variant"),
    (r"\b(humano|humana)\b",                   "human",      "standard"),
    (r"\bmeio[\s-]?orc\b",                     "half_orc",   None),
    (r"\bmeio[\s-]?elfo\b",                    "half_elf",   None),
    (r"\bhalfling\s+robust[oa]\b",             "halfling",   "stout"),
    (r"\bhalfling\s+pes?\s+leves?\b",          "halfling",   "lightfoot"),
    (r"\bhalfling\b",                          "halfling",   None),
    (r"\bpes?\s+macios?\b",                    "halfling",   "lightfoot"),
    (r"\banao?\s+da\s+colina\b",               "dwarf",      "hill"),
    (r"\banao?\s+colina\b",                    "dwarf",      "hill"),
    (r"\banao?\s+da\s+montanha\b",             "dwarf",      "mountain"),
    (r"\banao?\b",                             "dwarf",      None),
    (r"\balto\s+elfo\b",                       "elf",        "high"),
    (r"\bhigh\s+elf\b",                        "elf",        "high"),
    (r"\belfo\s+da\s+floresta\b",              "elf",        "wood"),
    (r"\bwood\s+elf\b",                        "elf",        "wood"),
    (r"\belfo\s+negro\b|\bdrow\b",             "elf",        "drow"),
    (r"\belfo\b",                              "elf",        None),
    (r"\baarakocra\s+cypsel\w*\b",             "aarakocra",  "cypselanos"),
    (r"\baaracokra\s+cypsel\w*\b",             "aarakocra",  "cypselanos"),
    (r"\baarakocra\s+falcon\w*\b",             "aarakocra",  "falconicos"),
    (r"\baarakocra\s+noctur\w*\b",             "aarakocra",  "nocturnos"),
    (r"\baarakocra\b|\baaracokra\b",           "aarakocra",  None),
    (r"\bgnomo\s+da\s+rocha\b",                "gnome",      "rock"),
    (r"\bgnomo\s+das\s+rochas\b",              "gnome",      "rock"),
    (r"\bgnomo\s+da\s+floresta\b",             "gnome",      "forest"),
    (r"\bgnomo\b",                             "gnome",      None),
    (r"\bminotauro\b",                         "minotaur",   None),
    (r"\bdraconato\b|\bdragonborn\b",          "dragonborn", None),
    (r"\btiefling\s+abissal\b",                "tiefling",   "abissal"),
    (r"\btiefling\s+ctonico\b",                "tiefling",   "ctonico"),
    (r"\btiefling\s+infernal\b",               "tiefling",   "infernal"),
    (r"\btiefling\b",                          "tiefling",   None),
    (r"\bcentauro\b",                          "centaur",    None),
]

# class palavra-chave (normalizada, sem acento) -> api_index
CLASS_KEYWORDS = [
    ("barbaro",      "barbarian"),
    ("barbara",      "barbarian"),
    ("barbado",      "barbarian"),    # erro de grafia em "Barbado do Caminho urso"
    ("bardo",        "bard"),
    ("barda",        "bard"),
    ("clerigo",      "cleric"),
    ("clerica",      "cleric"),
    ("druida",       "druid"),
    ("guerreiro",    "fighter"),
    ("guerreira",    "fighter"),
    ("monge",        "monk"),
    ("paladino",     "paladin"),
    ("paladina",     "paladin"),
    ("patrulheiro",  "ranger"),
    ("patrulheira",  "ranger"),
    ("ladino",       "rogue"),
    ("ladina",       "rogue"),
    ("feiticeiro",   "sorcerer"),
    ("feiticeira",   "sorcerer"),
    ("bruxo",        "warlock"),
    ("bruxa",        "warlock"),
    ("mago",         "wizard"),
    ("maga",         "wizard"),
    ("cozinheiro",   "cozinheiro"),
    ("cozinheirio",  "cozinheiro"),  # typo na ficha "Cozinheirio"
    ("cozinheira",   "cozinheiro"),
]

# Por classe, lista de (regex no resto-do-label, subclass_api_index)
#
# SUBCLASS_RULES: mapeia o LABEL bruto da planilha para o api_index REAL
# do DB (SubKlass.api_index), conforme api/docs/canonical_indexes.json.
# Quando ha duplicata hifen vs underscore no DB, preferimos a versao com
# hifen (formato mais novo/canonico).
#
SUBCLASS_RULES = {
    "barbarian": [
        (r"cicatrizes?\s+runicas?",                    "barbaro-cicatrizes-runicas"),
        (r"furios[oa]\s+imortal",                      "furioso-imortal"),
        (r"berserker|furios[oa](?!\s+imortal)",        "berserker"),
        (r"protetor\s+tribal",                         "protetor-tribal"),
        (r"raivoso\s+elemental",                       "raivoso-elemental"),
        (r"guerreiro\s+urso|urso",                     "guerreiro-urso"),
        (r"desistente",                                "desistente"),
        (r"totem|lobo|aguia",                          "totem"),
    ],
    "bard": [
        (r"conhecimento|lore",                         "lore"),
        (r"valor|bravura",                             "valor"),
        (r"comedia",                                   "colegio-comedia"),
        (r"fortuna",                                   "colegio-fortuna"),
        (r"pavor",                                     "colegio-pavor"),
        (r"quietude",                                  "colegio-quietude"),
        (r"virtuosismo",                               "colegio-virtuosismo"),
        (r"glamour",                                   "colegio-do-glamour"),
        (r"busca\s+(da\s+)?cancao",                    "colegio-busca-cancao"),
    ],
    "cleric": [
        (r"vida|life",                                 "dominio-da-vida"),
        (r"luz|light",                                 "dominio-da-luz"),
        (r"conhecimento|knowledge",                    "dominio-do-conhecimento"),
        (r"natureza|nature",                           "dominio-da-natureza"),
        (r"tempestade|tempest",                        "dominio-da-tempestade"),
        (r"trapaca|trickery|enganacao",                "dominio-da-trapaca"),
        (r"guerra|war",                                "dominio-da-guerra"),
        (r"agua",                                      "dominio-agua"),
        (r"\bar\b",                                    "dominio-ar"),
        (r"terra(?!\s+selvagem)",                      "dominio-terra"),
        (r"tempo",                                     "dominio-tempo"),
        (r"mente",                                     "dominio-mente"),
        (r"criacao",                                   "dominio-criacao"),
    ],
    "druid": [
        (r"circulo\s+da\s+terra|^terra$|land",         "circulo-da-terra"),
        (r"circulo\s+da\s+lua|^lua$|moon",             "circulo-da-lua"),
        (r"circulo\s+das\s+fadas|fadas",               "circulo-fadas"),
        (r"circulo\s+das\s+feras|feras",               "circulo-feras"),
        (r"circulo\s+(da\s+)?infestacao|infestacao",   "circulo-infestacao"),
        (r"circulo\s+(dos\s+)?mundos|mundos",          "circulo-mundos"),
        (r"circulo\s+(da\s+)?vida|vida",               "circulo-vida"),
        (r"circulo\s+verdejante|verdejante",           "verdejante"),
    ],
    "fighter": [
        (r"campeao|champion",                          "champion"),
        (r"mestre\s+de\s+batalha|battlemaster|battle\s+master", "mestre-de-batalha"),
        (r"cavaleiro\s+arcano|eldritch\s+knight|^arcano$|\barcano\b", "cavaleiro-arcano"),
        (r"atirador\s+inigualavel|gunslinger",         "atirador_inigualavel"),
        (r"cavaleiro\s+implacavel",                    "cavaleiro_implacavel"),
        (r"defensor\s+dedicado",                       "defensor_dedicado"),
        (r"kensai",                                    "kensai"),
        (r"mestre\s+do\s+arremesso",                   "mestre_arremesso"),
        (r"mestre\s+das\s+correntes",                  "mestre_correntes"),
    ],
    "monk": [
        (r"mao\s+aberta|open\s+hand",                  "mao-aberta"),
        (r"caminho\s+da\s+sombra|^sombras?$|shadow",   "sombra"),
        (r"quatro\s+elementos|four\s+elements",        "quatro-elementos"),
        (r"caminho\s+do\s+aco|^aco$",                  "caminho_aco"),
        (r"mestre\s+bebado|bebado|drunken",            "caminho_mestre_bebado"),
        (r"monge\s+tatuado|tatuado",                   "caminho_monge_tatuado"),
        (r"ninjut|ninjutsu",                           "caminho_ninjuts"),
        (r"punho\s+sagrado",                           "caminho_punho_sagrado"),
        (r"sadhaka",                                   "caminho_sadhaka"),
    ],
    "paladin": [
        (r"devocao|devotion",                          "devotion"),
        (r"anciao|anciaos|ancients|ancioes",           "ancients"),
        (r"vinganca|vengeance",                        "vengeance"),
        (r"misericordia|mercy",                        "juramento-misericordia"),
        (r"danacao",                                   "juramento-danacao"),
        (r"equilibrio",                                "juramento-equilibrio"),
        (r"liberdade",                                 "juramento-liberdade"),
        (r"ordenacao",                                 "juramento-ordenacao"),
        (r"pureza",                                    "juramento-pureza"),
    ],
    "ranger": [
        (r"cacador(?!\s+(de\s+)?tesouros?)|hunter",    "hunter"),
        (r"mestre\s+das?\s+(bestas|feras)|beast\s+master|beastmaster", "beast_master"),
        (r"flagelo\s+(dos\s+)?inimigos",               "flagelo-dos-inimigos"),
        (r"batedor|scout",                             "rastreador_urbano"),
        (r"arqueiro\s+(da\s+)?floresta",               "arqueiro_floresta_alta"),
        (r"guardiao\s+selvagem",                       "guardiao_selvagem"),
        (r"rastreador\s+urbano",                       "rastreador_urbano"),
    ],
    "rogue": [
        (r"ladrao|thief",                              "ladrao"),
        (r"assassin[oa]?|assasin[oa]?",                "assassino"),
        (r"trapaceiro\s+arcano|arcane\s+trickster",    "trapaceiro-arcano"),
        (r"cacador\s+(de\s+)?tesouros?|treasure",      "cacador-de-tesouros"),
        (r"dancarino\s+(das\s+)?sombras?",             "dancarino-das-sombras"),
        (r"face\s+fantasmagorica",                     "face-fantasmagorica"),
        (r"lamina\s+invisivel",                        "lamina-invisivel"),
        (r"larapio\s+(de\s+)?almas?",                  "larapio-de-almas"),
        (r"mimetizador",                               "mimetizador"),
    ],
    "sorcerer": [
        (r"draconic[oa]|draconic",                     "draconic"),
        (r"selvagem|wild",                             "wild"),
        (r"feiticaria\s+(da\s+)?espada|^espada$",      "feiticaria-da-espada"),
        (r"feiticaria\s+(do\s+)?sangue|^sangue$",      "feiticaria-do-sangue"),
        (r"linhagem\s+elemental",                      "linhagem-elemental"),
        (r"origem\s+aberrante",                        "origem-aberrante"),
        (r"origem\s+abissal",                          "origem-abissal"),
        (r"origem\s+mutavel",                          "origem-mutavel"),
    ],
    "warlock": [
        (r"infero|infernal|fiend|corruptor",           "fiend"),
        (r"arquefada|archfey|^fadas?$",                "archfey"),
        (r"antigo|great\s+old\s+one|grande\s+antigo",  "great_old_one"),
        (r"morte|death",                               "patrono-morte"),
        (r"arcanjo\s+vingador",                        "patrono-arcanjo-vingador"),
        (r"espirito\s+heroico",                        "patrono-espirito-heroico"),
        (r"supragenio",                                "patrono-supragenio"),
        (r"tita\s+caido",                              "patrono-tita-caido"),
        (r"^vazio$|patrono\s+vazio",                   "patrono-vazio"),
    ],
    "wizard": [
        (r"abjuracao",                                 "escola-de-abjuracao"),
        (r"conjuracao",                                "escola-de-conjuracao"),
        (r"adivinhacao",                               "escola-de-adivinhacao"),
        (r"encantamento",                              "escola-de-encantamento"),
        (r"evocacao",                                  "escola-de-evocacao"),
        (r"ilusao",                                    "escola-de-ilusao"),
        (r"necromancia",                               "escola-de-necromancia"),
        (r"transmutacao",                              "escola-de-transmutacao"),
        (r"navegacao\s+planar|^planar$",               "navegacao-planar"),
        (r"automatos|maestria\s+dos\s+automatos",      "maestria-dos-automatos"),
        (r"alquimica|alquimia",                        "maestria-alquimica"),
        (r"teurgia\s+mistica|teurgia",                 "teurgia-mistica"),
        (r"arquearia\s+arcana",                        "arquearia-arcana"),
        (r"iniciacao\s+(em\s+)?demonologia|demonologia", "iniciacao-demonologia"),
    ],
    "cozinheiro": [
        (r"mestre\s+da\s+fritura",                     "mestre-da-fritura"),
        (r"alquimista\s+gourmet",                      "alquimista-gourmet"),
        (r"doceiro\s+encantado",                       "doceiro-encantado"),
        (r"mestre\s+do\s+fogo\s+e\s+fumaca",           "mestre-do-fogo-e-fumaca"),
        (r"cantineiro\s+de\s+guerra",                  "cantineiro-de-guerra"),
    ],
}


def parse_race_from_name(raw_name):
    """Recebe o nome cru do personagem (ex: "Lyra El'Asah (Wood Elf)") e devolve
    {name_clean, race_raw, race_api_index, subrace_api_index}.

    Se nao houver parenteses, retorna name_clean = raw_name e race_* = None."""
    if not raw_name:
        return {"name_clean": None, "race_raw": None,
                "race_api_index": None, "subrace_api_index": None}

    raw = str(raw_name).strip()
    m = re.search(r"^(.*?)\s*\(([^)]+)\)\s*$", raw)
    if m:
        name_clean = m.group(1).strip()
        race_raw = m.group(2).strip()
    else:
        # Fallback: tentar extrair a raca do nome direto (ex: "Nikos Humano").
        # Olhamos apenas a metade DIREITA da string para reduzir falso positivo.
        name_clean = raw
        race_raw = None
        norm_full = norm(raw)
        for pat, race, sub in RACE_PATTERNS:
            mm = re.search(pat, norm_full)
            if mm:
                # Reconstroi o trecho cru a partir do match no normalizado.
                # Nao temos correspondencia 1:1 com acentos, entao usamos o
                # match no proprio norm_full como rotulo cru (suficiente).
                race_raw = mm.group(0)
                # Remove o trecho do nome para name_clean (best-effort, no raw)
                name_clean = re.sub(pat, "", norm_full).strip()
                # Restaura capitalizacao basica (so primeira letra)
                name_clean = name_clean.title() if name_clean else raw
                break

    if not race_raw:
        return {"name_clean": name_clean, "race_raw": None,
                "race_api_index": None, "subrace_api_index": None}

    race_norm = norm(race_raw)
    race_api = None
    subrace_api = None
    for pat, race, sub in RACE_PATTERNS:
        if re.search(pat, race_norm):
            race_api = race
            subrace_api = sub
            break

    return {
        "name_clean": name_clean,
        "race_raw": race_raw,
        "race_api_index": race_api,
        "subrace_api_index": subrace_api,
    }


def parse_class_from_label(raw_label):
    """Recebe a string de classe ("Patrulheiro Batedor", "Mago Planar",
    "Guerreira Mestre de Batalha", "Cozinheirio") e devolve mapeamento
    canonico + flag homebrew."""
    if not raw_label:
        return {
            "class_label_raw": None, "class_api_index": None,
            "subclass_label_raw": None, "subclass_api_index": None,
            "is_homebrew_class": False, "is_homebrew_subclass": False,
        }

    raw = str(raw_label).strip()
    norm_label = norm(raw)

    # Identifica a classe pela primeira palavra-chave que casar
    class_api = None
    matched_kw = None
    for kw, api in CLASS_KEYWORDS:
        if re.search(rf"\b{kw}\b", norm_label):
            class_api = api
            matched_kw = kw
            break

    # Resto do label = subclass label. Trabalhamos com a versao normalizada
    # (sem acentos) para evitar problemas tipo "Drúida" nao casar em \bdruida\b.
    subclass_label_raw = None
    if matched_kw:
        rest = re.sub(
            rf"\b{matched_kw}\b", "", norm_label, count=1
        ).strip(" -—()")
        rest = re.sub(r"^(do|da|de|dos|das)\s+", "", rest)
        subclass_label_raw = rest if rest else None
    else:
        subclass_label_raw = raw

    # Identifica a subclass canonica (so se identificamos a classe)
    subclass_api = None
    if class_api and subclass_label_raw:
        sub_norm = norm(subclass_label_raw)
        for pat, sub_api in SUBCLASS_RULES.get(class_api, []):
            if re.search(pat, sub_norm):
                subclass_api = sub_api
                break

    # Fallback: se NAO identificamos a classe, tenta inferir por subclass
    # (caso comum: ficha so escreve "Atirador Inigualavel" sem dizer
    # "Guerreiro Atirador Inigualavel"). Varre TODAS as subclasses.
    if class_api is None:
        full_norm = norm_label
        for cls, rules in SUBCLASS_RULES.items():
            for pat, sub_api in rules:
                if re.search(pat, full_norm):
                    class_api = cls
                    subclass_api = sub_api
                    subclass_label_raw = raw
                    break
            if class_api:
                break

    is_homebrew_class = class_api is None
    is_homebrew_subclass = (
        bool(subclass_label_raw) and not subclass_api and not is_homebrew_class
    )

    return {
        "class_label_raw":     raw,
        "class_api_index":     class_api,
        "subclass_label_raw":  subclass_label_raw,
        "subclass_api_index":  subclass_api,
        "is_homebrew_class":   is_homebrew_class,
        "is_homebrew_subclass": is_homebrew_subclass,
    }


# Posicoes fixas (linha, coluna) — validadas em todas as 39 fichas
ABILITY_ROWS = {
    "strength":     (4,  1, 5),   # label row, col_label, row_value
    "dexterity":    (6,  1, 7),
    "constitution": (8,  1, 9),
    "intelligence": (10, 1, 11),
    "wisdom":       (12, 1, 13),
    "charisma":     (14, 1, 15),
}

SKILL_ROWS = [
    ("acrobatics",         "Acrobacia",         "dex",  4),
    ("animal_handling",    "Adestrar Animais",  "wis",  5),
    ("arcana",             "Arcanismo",         "int",  6),
    ("athletics",          "Atletismo",         "str",  7),
    ("deception",          "Enganacao",         "cha",  8),
    ("stealth",            "Furtividade",       "dex",  9),
    ("history",            "Historia",          "int", 10),
    ("intimidation",       "Intimidacao",       "cha", 11),
    ("insight",            "Intuicao",          "wis", 12),
    ("investigation",      "Investigacao",      "int", 13),
    ("medicine",           "Medicina",          "wis", 14),
    ("nature",             "Natureza",          "int", 15),
    ("perception",         "Percepcao",         "wis", 16),
    ("performance",        "Performance",       "cha", 17),
    ("persuasion",         "Persuasao",         "cha", 18),
    ("sleight_of_hand",    "Prestidigitacao",   "dex", 19),
    ("religion",           "Religiao",          "int", 20),
    ("survival",           "Sobrevivencia",     "wis", 21),
]

SAVE_ROWS = [
    ("strength",     4),
    ("dexterity",    5),
    ("constitution", 6),
    ("wisdom",       7),
    ("intelligence", 8),
    ("charisma",     9),
]


def norm(s):
    """Normaliza string: remove acentos, lowercase, strip."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()


def get(ws, row, col):
    v = ws.cell(row, col).value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def to_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    m = re.search(r"-?\d+", s)
    return int(m.group(0)) if m else None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else None


def find_label(ws, *patterns, max_row=None, max_col=None):
    """Procura primeira celula cujo texto bate em qualquer dos patterns
    (regex case-insensitive, apos normalizacao)."""
    pats = [re.compile(p, re.IGNORECASE) for p in patterns]
    R = max_row or ws.max_row
    C = max_col or ws.max_column
    for r in range(1, R + 1):
        for c in range(1, C + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            txt = norm(v)
            for p in pats:
                if p.search(txt):
                    return (r, c)
    return None


def extract_meta(ws):
    """Header (linhas 1-2): nome, classe, xp, nivel, proficiencia, sobrecarga.

    Layout em (1-2, 10-13):
      R1C10="Sobrecarga"   R1C11=threshold_leve  R1C12=penalty_leve
      R2C10="Sobrecarga P" R2C11=threshold_pesa  R2C12=penalty_pesa  R2C13=Deslo.(m)
    """
    raw_name  = get(ws, 1, 1)
    raw_class = get(ws, 1, 5)
    race_info = parse_race_from_name(raw_name)
    class_info = parse_class_from_label(raw_class)

    # Override manual quando a raca nao esta documentada na xlsx
    tab = ws.title
    race_source = "xlsx"
    if tab in MANUAL_RACE_OVERRIDES:
        race_api, subrace_api = MANUAL_RACE_OVERRIDES[tab]
        race_info["race_api_index"] = race_api
        race_info["subrace_api_index"] = subrace_api
        race_info["race_raw"] = race_info["race_raw"] or "(manual)"
        race_source = "manual_override"

    # Override de subclasse para fichas onde a info nao foi capturada
    subclass_source = "xlsx"
    if tab in MANUAL_SUBCLASS_OVERRIDES and not class_info["subclass_api_index"]:
        class_info["subclass_api_index"] = MANUAL_SUBCLASS_OVERRIDES[tab]
        class_info["is_homebrew_subclass"] = False
        subclass_source = "manual_override"

    bg_block = None
    if tab in MANUAL_BACKGROUND_OVERRIDES:
        bg_block = {
            "id":     MANUAL_BACKGROUND_OVERRIDES[tab],
            "source": "manual_override",
        }

    return {
        "name_raw":          raw_name,
        "name":              race_info["name_clean"],
        "skip_audit":        tab in SKIP_FROM_AUDIT,
        "race": {
            "raw":               race_info["race_raw"],
            "race_api_index":    race_info["race_api_index"],
            "subrace_api_index": race_info["subrace_api_index"],
            "source":            race_source,
        },
        "klass": {
            "raw":                  class_info["class_label_raw"],
            "class_api_index":      class_info["class_api_index"],
            "subclass_raw":         class_info["subclass_label_raw"],
            "subclass_api_index":   class_info["subclass_api_index"],
            "subclass_source":      subclass_source,
            "is_homebrew_class":    class_info["is_homebrew_class"],
            "is_homebrew_subclass": class_info["is_homebrew_subclass"],
        },
        "background":        bg_block,
        "xp":                to_int(get(ws, 2, 7)),
        "level":             to_int(get(ws, 2, 8)),
        "proficiency_bonus": to_int(get(ws, 2, 9)),
        "encumbrance": {
            "light_threshold_kg": to_float(get(ws, 1, 11)),
            "light_dex_penalty":  to_int(get(ws, 1, 12)),
            "heavy_threshold_kg": to_float(get(ws, 2, 11)),
            "heavy_dex_penalty":  to_int(get(ws, 2, 12)),
        },
    }


def extract_encumbrance_pesada(ws, meta):
    """Fallback: se posicao fixa nao deu valor, procura label "Sobrecarga P."
    em outras posicoes (algumas fichas tem em (32-33, 1-3))."""
    if meta["encumbrance"]["heavy_threshold_kg"] is not None:
        return
    pos = find_label(ws, r"sobrecarga\s*p")
    if not pos:
        return
    r, c = pos
    meta["encumbrance"]["heavy_threshold_kg"] = to_float(get(ws, r, c + 1))
    meta["encumbrance"]["heavy_dex_penalty"]  = to_int(get(ws, r, c + 2))


def extract_abilities(ws):
    out = {}
    for key, (r_lbl, c_lbl, r_val) in ABILITY_ROWS.items():
        out[key] = {
            "score": to_float(get(ws, r_val, 1)),
            "mod":   to_int(get(ws, r_lbl, 2)),
        }
    return out


def extract_skills(ws):
    out = []
    for key, label, ability, row in SKILL_ROWS:
        # nome canonico em coluna 3 (com possivel sufixo "(Des)") + treino em col 4
        name_cell = get(ws, row, 3)
        out.append({
            "key": key,
            "label": label,
            "ability": ability,
            "training_hours": to_float(get(ws, row, 4)) or 0.0,
            "raw_label_in_sheet": name_cell,
        })
    return out


def extract_saves(ws):
    out = []
    for key, row in SAVE_ROWS:
        out.append({
            "ability": key,
            "training_hours": to_float(get(ws, row, 6)) or 0.0,
            "total_hours_pool": to_float(get(ws, row, 7)) or 0.0,
        })
    return out


def extract_additional_proficiencies(ws):
    """Bloco "Proficiencia Adcionais" comeca na linha 16 (col 1) ate linha
    ~22, com nome em col 1 e treino(h) em col 2. Filtra vazias."""
    out = []
    # Encontrar inicio do bloco
    pos = find_label(ws, r"proficiencia\s*ad")
    start_row = pos[0] + 1 if pos else 17
    for r in range(start_row, start_row + 12):
        name = get(ws, r, 1)
        hours = to_float(get(ws, r, 2))
        if not name and hours is None:
            continue
        if name:
            out.append({"name": str(name), "training_hours": hours or 0.0})
    return out


def extract_hit_points(ws):
    return {
        "total":   to_int(get(ws, 5, 8)),
        "current": to_int(get(ws, 5, 9)),
    }


def extract_combat_block(ws):
    out = {
        "ac":            to_int(get(ws, 6, 9)),
        "spell_save_dc": to_int(get(ws, 13, 9)),
        "spell_attack":  to_int(get(ws, 13, 11)),
        "speed_m":       to_float(get(ws, 2, 13)),
    }
    out["passive_perception"] = to_int(get(ws, 10, 9))
    out["passive_insight"]    = to_int(get(ws, 10, 11))
    return out


def extract_weight(ws):
    return {
        "max_personal_kg":     to_float(get(ws, 9, 8)),
        "current_personal_kg": to_float(get(ws, 9, 9)),
        "max_backpack_kg":     to_float(get(ws, 9, 10)),
        "current_backpack_kg": to_float(get(ws, 9, 11)),
    }


def extract_coins(ws):
    """Bloco "Pecas"/"Nº" — busca PC, PP, PO, PL como labels e pega celula
    a direita (offset +1 col, mesma linha)."""
    out = {"copper": None, "silver": None, "gold": None, "platinum": None}
    mapping = {"PC": "copper", "PP": "silver", "PO": "gold", "PL": "platinum"}
    for label, key in mapping.items():
        pos = find_label(ws, rf"^{label}$")
        if pos:
            r, c = pos
            out[key] = to_float(get(ws, r, c + 1))
    return out


def extract_folego(ws):
    """Bloco Folego: full + atual. Procura label "Folego" e usa offset
    conhecido (full em (lbl_row+2, lbl_col), atual em (lbl_row+2, lbl_col+1))."""
    pos = find_label(ws, r"^folego$")
    if not pos:
        return None
    r, c = pos
    return {
        "full":    to_float(get(ws, r + 2, c)),
        "current": to_float(get(ws, r + 2, c + 1)),
    }


def extract_hp_extra(ws):
    pos = find_label(ws, r"hp\s*extra")
    if not pos:
        return None
    r, c = pos
    # valor geralmente uma linha abaixo na mesma coluna ou +1
    candidates = [(r + 1, c), (r, c + 1), (r + 1, c + 1)]
    for rr, cc in candidates:
        v = to_float(get(ws, rr, cc))
        if v is not None:
            return v
    return None


def extract_languages(ws):
    """Bloco "Idiomas": coleta linhas abaixo na MESMA coluna, mas para no
    primeiro label/header de outro bloco para nao puxar 'Itens Sinc.',
    'Magias Preparas', 'Espacos de Magia', 'Nv 1-4', 'Nº', etc."""
    pos = find_label(ws, r"^idiomas$")
    if not pos:
        return []
    r, c = pos
    STOP_RE = re.compile(
        r"^(itens|magias|espac|nv\s|n[oº°]\s*$|preparad|max\s|maximo|"
        r"talent|aljava|munic|peso|atual|hp|ca|deslo|equipa|montado|"
        r"tipo\s+de|sentido|estilo|exaust|fol(e|ê)go|carga|sobrecarg|"
        r"surto|indomavel|pocao|forma|d\.\s*total|d\.\s*gast|treino|"
        r"\s*$)",
        re.IGNORECASE,
    )
    out = []
    for rr in range(r + 1, min(r + 12, ws.max_row + 1)):
        v = get(ws, rr, c)
        if v is None:
            break
        s = str(v).strip()
        if not s or STOP_RE.match(s):
            break
        # Canonicalizacao: typos + traducoes alternativas viram a forma
        # registrada em config/race_rules.yml.
        n_low = norm(s)
        s_clean = (s.replace("Halflling", "Halfling")
                    .replace("Druico", "Druidico")
                    .replace("Infenal", "Infernal")
                    .replace("comun", "Comum")
                    .replace("Mino", "Minotauro"))
        canon_map = {
            'abissal':   'Abyssal',
            'elfo':      'Élfico',
            'elfico':    'Élfico',
            'goblinico': 'Goblin',
            'goblin':    'Goblin',
            'anao':      'Anão',
            'gnomico':   'Gnômico',
            'draconico': 'Dracônico',
        }
        if n_low in canon_map:
            s_clean = canon_map[n_low]
        out.append(s_clean)
    return out


def extract_spell_slots(ws):
    """Procura por padroes "Nv 1", "Nv 2" ... "Nv 9" e captura quantidade
    a esquerda OU a direita."""
    out = []
    pat = re.compile(r"^Nv\s*(\d+)\b", re.IGNORECASE)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            m = pat.match(str(v).strip())
            if not m:
                continue
            level = int(m.group(1))
            # quantidade adjacente: tenta col+1 primeiro, depois col-1
            qty = to_float(get(ws, r, c + 1))
            if qty is None:
                qty = to_float(get(ws, r, c - 1))
            out.append({"level": level, "total": qty})
    # de-duplica por nivel (primeira ocorrencia ganha)
    seen = {}
    for s in out:
        if s["level"] not in seen and s["total"]:
            seen[s["level"]] = s
    return [seen[k] for k in sorted(seen.keys())]


def extract_synced_items(ws):
    """Itens Sincronizados — Max N + atual."""
    pos = find_label(ws, r"itens\s*sinc")
    if not pos:
        return None
    r, c = pos
    # "Máx 3" geralmente em (r+1, c) e atual em (r+1, c+1)
    raw_max = get(ws, r + 1, c)
    out = {"max": to_int(raw_max), "current": to_int(get(ws, r + 1, c + 1))}
    return out


def extract_class_resources(ws):
    """Captura recursos de classe quando os labels existem."""
    out = {}
    rules = [
        ("action_surge",   r"surto\s*de\s*a"),
        ("indomitable",    r"indom"),
        ("bardic_inspiration", r"inspira(c|ç)ão\s*b(a|á)rdica"),
        ("rage",           r"^f(u|ú)ria$"),
        ("ki",             r"^ki$"),
        ("sorcery_points", r"pontos\s*de\s*feiti"),
        ("channel_div",    r"canalizar\s*divindade"),
    ]
    for key, pat in rules:
        pos = find_label(ws, pat)
        if pos:
            r, c = pos
            out[key] = {
                "raw_label": get(ws, r, c),
                "value_right": to_float(get(ws, r, c + 1)),
                "value_below": to_float(get(ws, r + 1, c)),
            }
    return out


def extract_inventory(ws):
    """Bloco "Objetos na BAG" + colunas adjacentes (Quant, Peso). Captura ate
    encontrar linha completamente vazia."""
    pos = find_label(ws, r"objetos\s*na\s*bag")
    if not pos:
        return []
    r0, c0 = pos
    out = []
    blanks = 0
    for r in range(r0 + 2, ws.max_row + 1):
        name = get(ws, r, c0)
        qty  = to_float(get(ws, r, c0 + 1))
        wt   = to_float(get(ws, r, c0 + 2))
        if not name and qty is None and wt is None:
            blanks += 1
            if blanks >= 2:
                break
            continue
        blanks = 0
        if name:
            out.append({"name": str(name), "quantity": qty, "weight_each_kg": wt})
    return out


def extract_armor_weapons(ws):
    """Tabela "Armas/Peso" + "Vestindo/Peso" + "Amadura/Peso"."""
    out = {"armor": [], "weapons": [], "wearing": []}
    blocks = [
        ("armor",   r"^amadura$|^armadura$"),
        ("weapons", r"^armas$"),
        ("wearing", r"^vestindo$"),
    ]
    for key, pat in blocks:
        pos = find_label(ws, pat)
        if not pos:
            continue
        r0, c0 = pos
        for r in range(r0 + 1, r0 + 12):
            name = get(ws, r, c0)
            wt   = to_float(get(ws, r, c0 + 1))
            if not name and wt is None:
                break
            if name:
                out[key].append({"name": str(name), "weight_kg": wt})
    return out


def extract_spells_known(ws):
    """Lista de magias conhecidas/preparadas — comeca apos label "Magias" na
    coluna 12-13. Cada linha tem [nivel/Tru, nome, ?, ?, marcador 'x']."""
    pos = find_label(ws, r"^magias$")
    if not pos:
        return []
    r0, c0 = pos
    SKIP_NAMES = {"nome", "lvl", "nivel", "nível"}
    out = []
    blanks = 0
    for r in range(r0 + 1, ws.max_row + 1):
        first = get(ws, r, c0)
        name = get(ws, r, c0 + 1)
        if not first and not name:
            blanks += 1
            if blanks >= 2:
                break
            continue
        blanks = 0
        if not name:
            continue
        if str(name).strip().lower() in SKIP_NAMES:
            continue
        nivel_raw = first
        if isinstance(nivel_raw, str) and re.match(r"tru", nivel_raw, re.IGNORECASE):
            level = 0
        else:
            level = to_int(nivel_raw)
        marker = None
        for off in range(2, 6):
            v = get(ws, r, c0 + off)
            if v and str(v).strip().lower() in ("x", "sempre"):
                marker = str(v).strip().lower()
                break
        out.append({
            "level": level,
            "name": str(name),
            "marker": marker,
            "always_prepared": marker == "sempre",
        })
    return out


def extract_feats(ws):
    """Bloco "Talentos"/"Talento": lista de strings imediatamente abaixo do
    label, na MESMA coluna. Para de coletar quando encontra string com numero,
    label de outro bloco, ou celula vazia."""
    pos = find_label(ws, r"^talentos?$")
    if not pos:
        return []
    r0, c0 = pos
    STOP_RE = re.compile(
        r"^(exaust|espac|tipo|magias|nv\s|folego|aljava|munic|idiomas|peso|"
        r"atual|maximo|espaços|tipo de|hp|ca|deslo|bonus|bônus|treino|sentido)",
        re.IGNORECASE,
    )
    out = []
    for r in range(r0 + 1, min(r0 + 8, ws.max_row + 1)):
        v = get(ws, r, c0)
        if v is None:
            break
        s = str(v).strip()
        if not s or STOP_RE.match(s):
            break
        if re.search(r"\d", s) and len(s) <= 4:  # ex: "1.0", "5"
            break
        out.append(s)
    return out


def extract_fighting_style(ws):
    """Estilo de Luta: string na celula imediatamente abaixo do label."""
    pos = find_label(ws, r"estilo\s+de\s+luta")
    if not pos:
        return None
    r, c = pos
    v = get(ws, r + 1, c)
    return str(v) if v else None


def extract_ranger_choices(ws):
    """Inimigo Favorito + Terreno Favorito (Patrulheiro). Cada label tem o
    valor da escolha em (r+1, c). Pode aparecer multiplas vezes (nv 1+6+14)."""
    out = {"favored_enemy": [], "favored_terrain": []}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            n = norm(v)
            if "inimigo favorito" in n:
                val = get(ws, r + 1, c)
                if val:
                    out["favored_enemy"].append(str(val))
            elif "terreno favorito" in n:
                val = get(ws, r + 1, c)
                if val:
                    out["favored_terrain"].append(str(val))
    return out if (out["favored_enemy"] or out["favored_terrain"]) else None


def extract_rage(ws):
    """Furia (Barbaro): label "Furia"/"Fúria" + valor cru a direita.
    Formatos vistos: "1 de 3P", "2P", 1.0 — devolve string crua + parsed."""
    pos = find_label(ws, r"^f(u|ú)ria$")
    if not pos:
        return None
    r, c = pos
    raw = get(ws, r, c + 1)
    if raw is None:
        return None
    s = str(raw).strip()
    used, total = None, None
    # Padrao "X de YP"
    m = re.match(r"(\d+)\s*de\s*(\d+)\s*p?", s, re.IGNORECASE)
    if m:
        used = int(m.group(1))
        total = int(m.group(2))
    else:
        m = re.match(r"(\d+)\s*p", s, re.IGNORECASE)
        if m:
            total = int(m.group(1))
            used = 0
        else:
            try:
                total = int(float(s))
                used = 0
            except (ValueError, TypeError):
                pass
    return {"raw": s, "used": used, "total": total}


def extract_mount(ws):
    """Bloco "Equipamentos do Cavalo" + "Equipae" + "Montado" (Sim/Nao)
    + Carga total. Layout varia, capturamos generico."""
    pos = find_label(ws, r"equipamentos\s+do\s+cavalo")
    if not pos:
        # Fallback: procura "Equipae" sozinho (alguns layouts)
        pos = find_label(ws, r"^equipae$")
    if not pos:
        return None
    r0, c0 = pos
    # Procura "Equipae" header (pode estar na proxima linha)
    items = []
    header_row = None
    for r in range(r0, r0 + 3):
        v = get(ws, r, c0)
        if v and "equipae" in norm(v):
            header_row = r
            break
    if header_row is None:
        header_row = r0 + 1
    for r in range(header_row + 1, header_row + 12):
        name = get(ws, r, c0)
        qty  = to_float(get(ws, r, c0 + 1))
        wt   = to_float(get(ws, r, c0 + 2))
        if not name and qty is None:
            break
        if name:
            items.append({"name": str(name), "quantity": qty, "weight_each_kg": wt})

    # "Montado" (Sim/Nao) e "Carga" — varredura mais ampla nas colunas vizinhas
    mount_flag = None
    carga_total = None
    for r in range(r0, r0 + 14):
        for c in range(max(c0 - 1, 1), min(c0 + 8, ws.max_column + 1)):
            v = get(ws, r, c)
            if v is None:
                continue
            n = norm(v)
            if n == "montado":
                # Sim/Nao normalmente esta a direita e/ou na linha abaixo
                for (rr, cc) in [(r, c + 1), (r + 1, c + 1), (r + 1, c), (r + 2, c)]:
                    val = get(ws, rr, cc)
                    if isinstance(val, str) and val.strip().lower() in ("sim", "nao", "não"):
                        mount_flag = val.strip().lower() == "sim"
                        break
            elif n == "carga":
                # numero a direita ou abaixo
                for (rr, cc) in [(r, c + 1), (r + 1, c), (r + 1, c + 1)]:
                    cv = to_float(get(ws, rr, cc))
                    if cv is not None and cv > 0:
                        carga_total = cv
                        break

    if not items and mount_flag is None and carga_total is None:
        return None
    return {"items": items, "mounted": mount_flag, "load_kg": carga_total}


def extract_exhaustion(ws):
    """Exausto Nv / Exaustão Nv — valor numerico (0-6)."""
    pos = find_label(ws, r"exaust(ã|a)o\s+nv|exausto\s+nv")
    if not pos:
        return None
    r, c = pos
    v = to_int(get(ws, r, c + 1)) or to_int(get(ws, r + 1, c))
    return v


def extract_aljava(ws):
    pos = find_label(ws, r"^aljava$")
    if not pos:
        return None
    r, c = pos
    out = []
    for rr in range(r + 2, r + 12):
        name = get(ws, rr, c)
        qty  = to_float(get(ws, rr, c + 1))
        wt   = to_float(get(ws, rr, c + 2))
        if not name and qty is None:
            break
        if name:
            out.append({"name": str(name), "quantity": qty, "weight_each_kg": wt})
    return out


def extract_one(ws):
    sheet = {
        "tab_name": ws.title,
        "meta": extract_meta(ws),
        "abilities": extract_abilities(ws),
        "skills": extract_skills(ws),
        "saving_throws": extract_saves(ws),
        "additional_proficiencies": extract_additional_proficiencies(ws),
        "hit_points": extract_hit_points(ws),
        "hp_extra": extract_hp_extra(ws),
        "combat": extract_combat_block(ws),
        "weight": extract_weight(ws),
        "coins": extract_coins(ws),
        "folego": extract_folego(ws),
        "languages": extract_languages(ws),
        "spell_slots": extract_spell_slots(ws),
        "spells_listed": extract_spells_known(ws),
        "synced_items": extract_synced_items(ws),
        "aljava": extract_aljava(ws),
        "inventory_bag": extract_inventory(ws),
        "armor_weapons": extract_armor_weapons(ws),
        "class_resources_signals": extract_class_resources(ws),
        "feats": extract_feats(ws),
        "fighting_style": extract_fighting_style(ws),
        "ranger_choices": extract_ranger_choices(ws),
        "rage": extract_rage(ws),
        "mount": extract_mount(ws),
        "exhaustion_level": extract_exhaustion(ws),
    }
    extract_encumbrance_pesada(ws, sheet["meta"])
    return sheet


def main():
    if not SRC_XLSX.exists():
        print(f"ERRO: nao achei {SRC_XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    sheets = []
    for tab in wb.sheetnames:
        if tab in ABAS_IGNORAR:
            continue
        ws = wb[tab]
        try:
            sheets.append(extract_one(ws))
        except Exception as e:
            print(f"ERRO extraindo {tab!r}: {type(e).__name__}: {e}", file=sys.stderr)

    OUT_JSON.write_text(json.dumps(sheets, ensure_ascii=False, indent=2), encoding="utf-8")

    # Relatorio resumido
    lines = [f"Extracao: {len(sheets)} fichas\n", "=" * 90, ""]
    fmt = (
        "{tab:<12} | {name:<25} | "
        "{race:<14} | {klass:<10} | {sub:<18} | nv {lvl:<2} | "
        "HP {hp_c}/{hp_t} | CA {ac} | sp {sp:>2} | inv {inv:>2}"
    )

    def fmt_klass(k):
        if not k:
            return "?", "?"
        if k["class_api_index"]:
            cls = k["class_api_index"]
        else:
            cls = f"~{(k['raw'] or '?')[:8]}"
        if k["subclass_api_index"]:
            sub = k["subclass_api_index"]
        elif k["subclass_raw"]:
            sub = f"~{k['subclass_raw'][:16]}"
        else:
            sub = "-"
        if k["is_homebrew_class"]:
            cls = "!" + cls
        if k["is_homebrew_subclass"]:
            sub = "!" + sub
        return cls, sub

    def fmt_race(r):
        if not r or not r["race_api_index"]:
            return f"!{(r['raw'] or '?')[:12]}" if r and r["raw"] else "?"
        if r["subrace_api_index"]:
            return f"{r['race_api_index']}/{r['subrace_api_index']}"
        return r["race_api_index"]

    for s in sheets:
        m = s["meta"]
        hp = s["hit_points"]
        cb = s["combat"]
        cls, sub = fmt_klass(m.get("klass"))
        lines.append(fmt.format(
            tab=s["tab_name"],
            name=(m["name"] or m["name_raw"] or "?")[:25],
            race=fmt_race(m.get("race"))[:14],
            klass=cls[:10],
            sub=sub[:18],
            lvl=m["level"],
            hp_c=hp["current"], hp_t=hp["total"], ac=cb["ac"],
            sp=len(s["spells_listed"]), inv=len(s["inventory_bag"]),
        ))
    lines.append("")
    lines.append("Legenda: ! = homebrew (sem mapeamento canonico no projeto)")
    lines.append("         ~xxx = label cru (nao mapeado)")
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")

    print(f"OK: {len(sheets)} fichas escritas em {OUT_JSON}")
    print(f"    relatorio em {OUT_TXT}")


if __name__ == "__main__":
    main()
